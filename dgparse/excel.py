#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Centralized Parser for models coming from Microsoft Excel Spreadsheets
"""

from functools import partial
import logging

import openpyxl

from .sequtils import dotsetter

log = logging.getLogger(__file__)


def row_to_dict(headers, constants, row_data):
    """Convert a row to a valid dictionary"""
    record = {}
    record.update(constants)
    if not row_data:  # handle dead rows
        return
    for key, value in zip(headers, row_data):
        if key.value is None:
            continue
        if hasattr(key, 'value'):
            key = key.value
        if hasattr(value, 'value') and value.value is not None:
            value = value.value
        else:
            continue
        if '.' in key:
            tokens = key.split('.')
            tokens.reverse()
            dotsetter(tokens, value, record)
        else:
            record[key] = value
    return record

def parse(open_file):
    """Constructor which returns an excel parser callable for a given model"""
    records = []  # Use empty list to represent an empty set
    wbook = openpyxl.load_workbook(open_file, data_only=True)  #turn zipped file into workbook
    sheets = wbook.get_sheet_names()  # get the types of records included
    for sheet in sheets:
        wsheet = wbook[sheet]  # name of sheet is always the record type
        if not wsheet.rows or wsheet.max_row < 1:
            continue
        headers = wsheet[1]  # always the attributes
        record_factory = partial(row_to_dict, headers, {})  # no constants yet
        # This is dumb, but simpler than the broken comprehension
        for row in wsheet.iter_rows(min_row=2):
            if row:
                record = record_factory(row)
                if record:
                    records.append(record)
    return records
